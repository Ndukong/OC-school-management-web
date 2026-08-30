from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Exists, OuterRef, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from core.forms import StudentForm
from core.models import (
    AcademicTerm,
    SchoolClass,
    Student,
    StudentEnrollment,
    SubjectAverage,
    TermResult,
)
from core.utils.permissions import (
    get_school_for_user,
    is_admin_or_superuser,
    role_required,
)
from core.utils.tenancy import tenant_at_student_limit


def _apply_student_filters(qs, q="", class_id="", sex=""):
    """Apply search/class/sex filters. Search matches every word (token) of q
    anywhere in the student's name, ID, or guardian name — so both
    ``ADAMU`` and ``ADAMU NOELA`` find ``ADAMU NOELA NSHINEH``."""
    q = (q or "").strip()
    if q:
        token_filter = Q()
        for token in q.split():
            token_filter &= (
                Q(first_name__icontains=token)
                | Q(other_names__icontains=token)
                | Q(unique_id__icontains=token)
                | Q(guardian_name__icontains=token)
            )
        qs = qs.filter(token_filter)

    if class_id and str(class_id).isdigit():
        qs = qs.filter(
            Exists(
                StudentEnrollment.objects.filter(
                    student=OuterRef("pk"),
                    school_class_id=int(class_id),
                    academic_term__is_current=True,
                )
            )
        )

    if sex in ("M", "F"):
        qs = qs.filter(sex=sex)

    return qs


@login_required
@role_required("admin", "superuser")
def student_list(request):
    school = get_school_for_user(request.user)
    if not school:
        messages.error(request, "No school linked to your account.")
        return render(request, "students/list.html", {"students": [], "school": None})

    qs = Student.objects.filter(school=school, is_active=True)

    q = request.GET.get("q", "").strip()
    class_id = request.GET.get("class", "").strip()
    sex = request.GET.get("sex", "").strip()
    if class_id and not class_id.isdigit():
        class_id = ""
    if sex not in ("M", "F"):
        sex = ""

    qs = _apply_student_filters(qs, q=q, class_id=class_id, sex=sex)

    qs = qs.order_by("first_name", "other_names")

    # Per-page selector (default 25)
    try:
        per_page = int(request.GET.get("per_page", 25))
    except (TypeError, ValueError):
        per_page = 25
    if per_page not in (10, 25, 50, 100):
        per_page = 25

    paginator = Paginator(qs, per_page)

    # Guard against out-of-range / malformed page numbers
    try:
        page_num = int(request.GET.get("page", 1))
    except (TypeError, ValueError):
        page_num = 1
    page = paginator.get_page(page_num)

    classes = SchoolClass.objects.filter(school=school).order_by("sort_order")

    current_term = AcademicTerm.objects.filter(school=school, is_current=True).first()

    # Querystring fragment reused by pagination links (excludes page & per_page)
    qs_parts = []
    if q:
        qs_parts.append(f"q={q}")
    if class_id:
        qs_parts.append(f"class={class_id}")
    if sex:
        qs_parts.append(f"sex={sex}")
    page_qs = "&".join(qs_parts)
    if page_qs:
        page_qs += "&"

    ctx = {
        "page_obj": page,
        "students": page.object_list,
        "query": q,
        "class_filter": class_id,
        "sex_filter": sex,
        "per_page": per_page,
        "classes": classes,
        "current_term": current_term,
        "total_count": paginator.count,
        "page_qs": page_qs,
        "is_htmx": bool(request.headers.get("HX-Request")),
    }

    template = "students/_list_results.html" if ctx["is_htmx"] else "students/list.html"
    return render(request, template, ctx)


@login_required
@role_required("admin", "superuser")
def student_create(request):
    school = get_school_for_user(request.user)
    if not school:
        messages.error(request, "No school linked to your account.")
        return redirect("student_list")

    if request.method == "POST":
        if tenant_at_student_limit(school):
            messages.error(
                request,
                "Student limit reached for your current license. "
                "Contact the platform provider to increase your quota.",
            )
            return redirect("student_list")
        form = StudentForm(request.POST, request.FILES, school=school)
        if form.is_valid():
            student = form.save(commit=False)
            student.school = school
            student.save()
            messages.success(request, f"Student {student.full_name} registered.")
            return redirect("student_detail", pk=student.pk)
    else:
        form = StudentForm(school=school)

    return render(request, "students/create.html", {"form": form})


@login_required
def student_detail(request, pk):
    school = get_school_for_user(request.user)
    student = get_object_or_404(Student, pk=pk)
    if school is None or student.school_id != school.pk:
        messages.error(request, "Student not found.")
        return redirect("student_list")

    current_term = AcademicTerm.objects.filter(
        school=student.school, is_current=True
    ).first()
    enrollments = (
        StudentEnrollment.objects.filter(student=student)
        .select_related("school_class", "academic_term")
        .order_by("-academic_term__year_start", "-academic_term__term_number")
    )

    subject_averages = []
    term_result = None
    if current_term:
        subject_averages = SubjectAverage.objects.filter(
            student=student, academic_term=current_term
        ).select_related("subject")
        term_result = TermResult.objects.filter(
            student=student, academic_term=current_term
        ).first()

    return render(
        request,
        "students/detail.html",
        {
            "student": student,
            "enrollments": enrollments,
            "current_term": current_term,
            "subject_averages": subject_averages,
            "term_result": term_result,
            "is_admin": is_admin_or_superuser(request.user),
        },
    )


@login_required
@role_required("admin", "superuser")
def student_edit(request, pk):
    school = get_school_for_user(request.user)
    student = get_object_or_404(Student, pk=pk)
    if school is None or student.school_id != school.pk:
        messages.error(request, "Student not found.")
        return redirect("student_list")

    if request.method == "POST":
        form = StudentForm(request.POST, request.FILES, instance=student, school=school)
        if form.is_valid():
            form.save()
            messages.success(request, f"Student {student.full_name} updated.")
            return redirect("student_detail", pk=student.pk)
    else:
        form = StudentForm(instance=student, school=school)

    return render(request, "students/edit.html", {"form": form, "student": student})


@login_required
@role_required("admin", "superuser")
def student_export_excel(request):
    """Export filtered student list to .xlsx."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    school = get_school_for_user(request.user)
    if not school:
        return HttpResponse("No school", status=400)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )

    headers = [
        "#",
        "Name",
        "Sex",
        "Student ID",
        "Class",
        "DOB",
        "Guardian",
        "Contact",
        "Status",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    students = (
        _apply_student_filters(
            Student.objects.filter(school=school, is_active=True),
            q=request.GET.get("q", "").strip(),
            class_id=request.GET.get("class", "").strip(),
            sex=request.GET.get("sex", "").strip(),
        )
        .select_related()
        .order_by("first_name")
    )
    for row, s in enumerate(students, 2):
        enrollment = (
            s.enrollments.filter(academic_term__is_current=True)
            .select_related("school_class")
            .first()
        )
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=s.full_name)
        ws.cell(row=row, column=3, value=s.get_sex_display())
        ws.cell(row=row, column=4, value=s.unique_id)
        ws.cell(
            row=row, column=5, value=str(enrollment.school_class) if enrollment else ""
        )
        ws.cell(row=row, column=6, value=str(s.date_of_birth))
        ws.cell(row=row, column=7, value=s.guardian_name)
        ws.cell(row=row, column=8, value=s.guardian_contact)
        ws.cell(row=row, column=9, value="Active" if s.is_active else "Inactive")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 16

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="students_{school.matricule}.xlsx"'
    )
    wb.save(response)
    return response
