from datetime import datetime

from django.contrib import messages
from django.shortcuts import redirect, render
from openpyxl import load_workbook

from core.forms import StudentImportForm
from core.models import AcademicTerm, School, SchoolClass, Student, StudentEnrollment
from core.utils.permissions import get_school_for_user, role_required
from core.utils.tenancy import tenant_student_slots_remaining

PREVIEW_SESSION_KEY = "student_import_preview"

STATUS_NEW = "new"
STATUS_EXISTING = "existing"
STATUS_SKIPPED = "skipped"


def _parse_name(full_name: str) -> tuple:
    parts = full_name.strip().split(None, 1)
    return (parts[0] if parts else "", parts[1] if len(parts) > 1 else "")


def _parse_date(date_str: str):
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(date_str.strip(), fmt).date()
        except (ValueError, AttributeError):
            continue
    return None


def _detect_columns(rows) -> tuple:
    """Return (header_idx, col_map) or (None, {}) when no header row is found."""
    header_idx = None
    for i, row in enumerate(rows):
        vals = [str(v).strip().lower() if v else "" for v in row]
        if any("register" in v for v in vals):
            header_idx = i
            break
    if header_idx is None:
        return None, {}

    headers = [str(v).strip().lower() if v else "" for v in rows[header_idx]]
    col_map = {}
    for i, h in enumerate(headers):
        if "name" in h or "surname" in h:
            col_map["name"] = i
        elif "register" in h or "matric" in h or "id" in h:
            col_map["register"] = i
        elif h in ("sex", "gender"):
            col_map["sex"] = i
        elif "date" in h and "birth" in h:
            col_map["dob"] = i
        elif "place" in h and "birth" in h:
            col_map["pob"] = i
        elif "class" in h and "sub" not in h:
            col_map["class"] = i
        elif "subclass" in h or "sub" in h:
            col_map["subclass"] = i
        elif "subdivision" in h or "sub-division" in h or "arrondissement" in h:
            col_map["subdivision"] = i
    return header_idx, col_map


def _parse_sheet(rows, school: School, term: AcademicTerm) -> dict:
    """Validate workbook rows and return preview records without committing.

    Each record is JSON-safe (date as ISO string, class as id) so it can be
    stored in the session between preview and confirmation.
    """
    header_idx, col_map = _detect_columns(rows)
    if header_idx is None:
        return {"error": "Could not find a header row with 'Register'."}

    required = ["name", "register", "sex", "dob"]
    missing = [r for r in required if r not in col_map]
    if missing:
        headers = [str(v).strip().lower() if v else "" for v in rows[header_idx]]
        return {
            "error": f"Could not find columns: {', '.join(missing)}. Headers: {headers}"
        }

    records = []
    seen_registers = set()
    for i, row in enumerate(rows[header_idx + 1 :]):
        vals = list(row)
        rec = {
            "row": header_idx + 2 + i,
            "status": STATUS_NEW,
            "error": "",
            "note": "",
            "first_name": "",
            "other_names": "",
            "register": "",
            "sex": "",
            "dob": "",
            "pob": "",
            "subdivision": "",
            "school_class_id": None,
            "class_display": "",
        }

        name_raw = (
            str(vals[col_map["name"]]).strip() if col_map["name"] < len(vals) else ""
        )
        reg_raw = (
            str(vals[col_map["register"]]).strip()
            if col_map["register"] < len(vals)
            else ""
        )
        sex_raw = (
            str(vals[col_map["sex"]]).strip().upper()
            if col_map["sex"] < len(vals)
            else ""
        )
        dob_raw = (
            str(vals[col_map["dob"]]).strip() if col_map["dob"] < len(vals) else ""
        )

        rec["register"] = reg_raw
        if not name_raw or not reg_raw:
            rec["status"] = STATUS_SKIPPED
            rec["error"] = "Missing name or register number"
            records.append(rec)
            continue
        if reg_raw in seen_registers:
            rec["status"] = STATUS_SKIPPED
            rec["error"] = "Duplicate register number in file"
            records.append(rec)
            continue
        seen_registers.add(reg_raw)

        if sex_raw not in ("M", "F"):
            rec["status"] = STATUS_SKIPPED
            rec["error"] = f"Invalid sex '{sex_raw}' (expected M or F)"
            records.append(rec)
            continue

        dob = _parse_date(dob_raw)
        if not dob:
            rec["status"] = STATUS_SKIPPED
            rec["error"] = f"Invalid date of birth '{dob_raw}'"
            records.append(rec)
            continue

        first_name, other_names = _parse_name(name_raw)
        rec["first_name"] = first_name
        rec["other_names"] = other_names
        rec["sex"] = sex_raw
        rec["dob"] = dob.isoformat()
        rec["pob"] = (
            str(vals[col_map["pob"]]).strip()
            if col_map.get("pob", -1) < len(vals) and col_map.get("pob") is not None
            else ""
        )
        rec["subdivision"] = (
            str(vals[col_map["subdivision"]]).strip()
            if col_map.get("subdivision", -1) < len(vals)
            and col_map.get("subdivision") is not None
            else ""
        )

        class_name = (
            str(vals[col_map["class"]]).strip()
            if col_map.get("class", -1) < len(vals) and col_map.get("class") is not None
            else ""
        )
        rec["class_display"] = class_name
        if class_name:
            school_class = SchoolClass.objects.filter(
                school=school, name__iexact=class_name
            ).first()
            if not school_class:
                school_class = SchoolClass.objects.filter(
                    school=school, code__iexact=class_name[:6]
                ).first()
            if school_class:
                rec["school_class_id"] = school_class.id
                rec["class_display"] = str(school_class)
            else:
                rec["note"] = (
                    f"Class '{class_name}' not found — student imported without enrollment"
                )

        if Student.objects.filter(school=school, unique_id=reg_raw).exists():
            rec["status"] = STATUS_EXISTING

        records.append(rec)

    return {
        "header_idx": header_idx,
        "records": records,
        "new_count": len([r for r in records if r["status"] == STATUS_NEW]),
        "existing_count": len([r for r in records if r["status"] == STATUS_EXISTING]),
        "skipped_count": len([r for r in records if r["status"] == STATUS_SKIPPED]),
    }


def _commit_records(
    records: list,
    school: School,
    term: AcademicTerm,
    max_new: int | None = None,
) -> tuple:
    """Create students/enrollments, capped at ``max_new`` new students."""
    created = 0
    enrolled = 0
    quota_blocked = 0
    for rec in records:
        if rec["status"] == STATUS_SKIPPED:
            continue
        if rec["status"] == STATUS_NEW and max_new is not None and created >= max_new:
            quota_blocked += 1
            continue
        student, is_new = Student.objects.get_or_create(
            school=school,
            unique_id=rec["register"],
            defaults={
                "first_name": rec["first_name"],
                "other_names": rec["other_names"],
                "sex": rec["sex"],
                "date_of_birth": datetime.fromisoformat(rec["dob"]).date(),
                "place_of_birth": rec["pob"],
                "guardian_name": "",
                "division_of_origin": "",
                "sub_division_of_origin": rec["subdivision"],
                "region_of_origin": "",
            },
        )
        if is_new:
            created += 1

        school_class = None
        if rec["school_class_id"]:
            school_class = SchoolClass.objects.filter(
                pk=rec["school_class_id"], school=school
            ).first()
        if school_class:
            _, enr_new = StudentEnrollment.objects.get_or_create(
                student=student,
                academic_term=term,
                defaults={"school_class": school_class},
            )
            if enr_new:
                enrolled += 1
    return created, enrolled, quota_blocked


@role_required("admin")
def import_students_view(request):
    if request.method == "POST":
        if request.POST.get("confirm"):
            data = request.session.pop(PREVIEW_SESSION_KEY, None)
            if not data:
                messages.error(
                    request, "Preview expired — please upload the file again."
                )
                return redirect("import_students")
            school = (
                get_school_for_user(request.user)
                or School.objects.get(pk=data["school_id"])
            )
            term = AcademicTerm.objects.get(pk=data["term_id"])
            created, enrolled, quota_blocked = _commit_records(
                data["records"],
                school,
                term,
                tenant_student_slots_remaining(school),
            )
            if quota_blocked:
                messages.warning(
                    request,
                    f"License student limit reached — {quota_blocked} "
                    f"student(s) were not imported.",
                )
            messages.success(
                request,
                f"Import complete. Created {created} new students, "
                f"{enrolled} new enrollments. You can now edit their details.",
            )
            return redirect("student_list")

        form = StudentImportForm(request.POST, request.FILES)
        if form.is_valid():
            profile_school = get_school_for_user(request.user)
            school_id = int(form.cleaned_data["school"])
            if profile_school:
                school = profile_school
            else:
                school = School.objects.get(pk=school_id)
            term_id = form.cleaned_data.get("term")
            term = None
            if term_id:
                term = AcademicTerm.objects.filter(
                    pk=int(term_id), school=school
                ).first()
            if not term:
                term = AcademicTerm.objects.filter(
                    school=school, is_current=True
                ).first()
            if not term:
                term = AcademicTerm.objects.filter(school=school).first()
            if not term:
                messages.error(request, "No academic term found. Create one first.")
                return redirect("import_students")

            xlsx_file = request.FILES["file"]
            xlsx_file.seek(0)
            wb = load_workbook(xlsx_file, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))

            parsed = _parse_sheet(rows, school, term)
            if "error" in parsed:
                messages.error(request, parsed["error"])
                return redirect("import_students")

            request.session[PREVIEW_SESSION_KEY] = {
                "school_id": school.id,
                "term_id": term.id,
                "file_name": xlsx_file.name,
                "records": parsed["records"],
            }

            stats = {
                "students": Student.objects.filter(school=school).count(),
                "enrollments": StudentEnrollment.objects.filter(
                    student__school=school
                ).count(),
            }
            return render(
                request,
                "students/import.html",
                {
                    "form": StudentImportForm(),
                    "stats": stats,
                    "preview": parsed,
                    "school": school,
                    "term": term,
                    "file_name": xlsx_file.name,
                },
            )
    else:
        if request.GET.get("cancel"):
            request.session.pop(PREVIEW_SESSION_KEY, None)
            return redirect("import_students")
        form = StudentImportForm()

    profile_school = get_school_for_user(request.user)
    if profile_school:
        stats = {
            "students": Student.objects.filter(school=profile_school).count(),
            "enrollments": StudentEnrollment.objects.filter(
                student__school=profile_school
            ).count(),
        }
    else:
        stats = {
            "students": Student.objects.count(),
            "enrollments": StudentEnrollment.objects.count(),
        }
    return render(request, "students/import.html", {"form": form, "stats": stats})
