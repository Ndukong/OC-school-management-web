"""Import official Cameroon GCE subjects and competencies from the standard xlsx file.

Usage:
    python manage.py import_xlsx_defaults                       # import both subjects + competencies
    python manage.py import_xlsx_defaults --subjects-only       # subjects only
    python manage.py import_xlsx_defaults --competencies-only   # competencies only
    python manage.py import_xlsx_defaults --auto               # skip prompts

The command is idempotent — re-running won't create duplicates.
"""

import os

from django.conf import settings
from django.core.management.base import BaseCommand

XLSX_PATH = os.path.join(settings.BASE_DIR, "Subjects and comptencies.xlsx")

TERM_MAP = {"FIRST": 1, "SECOND": 2, "THIRD": 3}


class Command(BaseCommand):
    help = (
        "Import subjects and first-cycle competencies from the official Cameroon xlsx."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--auto", action="store_true", help="Skip confirmation prompt"
        )
        parser.add_argument(
            "--subjects-only",
            action="store_true",
            help="Import only subjects (skip competencies)",
        )
        parser.add_argument(
            "--competencies-only",
            action="store_true",
            help="Import only competencies (skip subjects)",
        )
        parser.add_argument(
            "--school",
            type=str,
            default=None,
            help="Import for the school with this pk (multi-tenant).",
        )

    def handle(self, *args, **options):
        from core.models import School

        if not os.path.exists(XLSX_PATH):
            self.stderr.write(self.style.ERROR(f"File not found: {XLSX_PATH}"))
            return

        try:
            import openpyxl
        except ImportError:
            self.stderr.write(
                self.style.ERROR("openpyxl not installed: pip install openpyxl")
            )
            return

        if options["school"]:
            school = School.objects.filter(pk=options["school"]).first()
        else:
            school = School.objects.filter(is_active=True).first()
        if not school:
            self.stderr.write(
                self.style.ERROR(
                    "No school found. Run the activation wizard first."
                )
            )
            return

        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

        import_subjects = not options["competencies_only"]
        import_competencies = not options["subjects_only"]

        if import_subjects:
            self._import_subjects(wb, school)
        if import_competencies:
            self._import_competencies(wb, school)

        self.stdout.write(self.style.SUCCESS("Done."))

    # ── Subjects ────────────────────────────────────────────────────────────

    def _import_subjects(self, wb, school):
        from core.models import Subject

        ws = wb["Subjects"]
        created = 0
        skipped = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            code = (r[2] or "").strip()
            name = (r[3] or "").strip()
            order = int(r[4]) if r[4] else 0
            if not code or not name:
                continue
            code = code.upper()
            _, was_created = Subject.objects.get_or_create(
                school=school,
                code=code,
                defaults={"name": name, "sort_order": order},
            )
            if was_created:
                created += 1
            else:
                skipped += 1
        self.stdout.write(
            self.style.SUCCESS(
                f"  Subjects: {created} created, {skipped} already existed"
            )
        )

    # ── Competencies ────────────────────────────────────────────────────────

    def _import_competencies(self, wb, school):
        from core.models import AcademicTerm, Competency, Subject

        ws = wb["Competencies"]

        # Build name → Subject map (case-insensitive, with known aliases)
        NAME_ALIASES = {
            "additional maths": "ama",
            "additional math": "ama",
            "sports and physical education": "spo",
            "sports & physical ed.": "spo",
            "sports & physical ed": "spo",
            "food & nutrition": "fnu",
            "food and nutrition": "fnu",
            "human biology": "hbi",
            "computer sciences": "csc",
            "computer science": "csc",
        }
        subject_map = {}
        for subj in Subject.objects.filter(school=school):
            subject_map[subj.name.lower()] = subj
            subject_map[subj.code.lower()] = subj
        for alias, code in NAME_ALIASES.items():
            if alias not in subject_map and code in subject_map:
                subject_map[alias] = subject_map[code]

        # Build term_number → AcademicTerm map (use current year terms)
        term_map = {}
        for term in AcademicTerm.objects.filter(school=school).order_by(
            "year_start", "term_number"
        ):
            term_map[term.term_number] = term

        created = 0
        skipped = 0
        warnings = []
        seen = set()

        for r in ws.iter_rows(min_row=3, values_only=True):
            term_str = (r[1] or "").strip().upper()
            subj_str = (r[2] or "").strip()

            if term_str not in TERM_MAP:
                continue
            if not subj_str or subj_str.lower() == "subject":
                continue

            term_number = TERM_MAP[term_str]
            term = term_map.get(term_number)
            if not term:
                continue

            subject = subject_map.get(subj_str.lower())
            if not subject:
                warnings.append(
                    f"Subject '{subj_str}' not found — skipping competencies"
                )
                continue

            # Extract up to 4 competencies from columns D-G
            comp_texts = []
            for col_idx in range(3, 7):
                val = r[col_idx]
                if val and str(val).strip() and str(val).strip() != '"':
                    text = str(val).strip()
                    if text not in comp_texts:
                        comp_texts.append(text)

            if not comp_texts:
                continue

            for sort_order, description in enumerate(comp_texts, start=1):
                # Create for form_levels 1–5 (all first-cycle forms share competencies)
                for form_level in range(1, 6):
                    key = (subject.pk, term.pk, form_level, sort_order)
                    if key in seen:
                        continue
                    seen.add(key)

                    _, was_created = Competency.objects.get_or_create(
                        subject=subject,
                        term=term,
                        form_level=form_level,
                        sort_order=sort_order,
                        defaults={"description": description},
                    )
                    if was_created:
                        created += 1
                    else:
                        skipped += 1

        if warnings:
            seen_warns = set()
            for w in warnings:
                if w not in seen_warns:
                    self.stdout.write(self.style.WARNING(f"  {w}"))
                    seen_warns.add(w)

        self.stdout.write(
            self.style.SUCCESS(
                f"  Competencies: {created} created, {skipped} already existed "
                f"(term × subject × form_level × sort_order)"
            )
        )
