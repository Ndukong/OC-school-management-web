from datetime import datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from core.models import AcademicTerm, School, SchoolClass, Student, StudentEnrollment


class Command(BaseCommand):
    help = "Import students from the government .xlsx list"

    def add_arguments(self, parser):
        parser.add_argument("--school", type=int, required=True, help="School ID")
        parser.add_argument(
            "--file", type=str, required=True, help="Path to .xlsx file"
        )
        parser.add_argument(
            "--term", type=int, help="Term ID (defaults to current term)"
        )
        parser.add_argument("--year-start", type=int, default=2025)
        parser.add_argument("--year-end", type=int, default=2026)

    def _parse_name(self, full_name: str) -> tuple:
        """Split 'ADAMU SALE' into first_name='ADAMU', other_names='SALE'."""
        parts = full_name.strip().split(None, 1)
        first = parts[0] if parts else ""
        rest = parts[1] if len(parts) > 1 else ""
        return first, rest

    def _parse_date(self, date_str: str):
        """Parse DD/MM/YYYY to date object."""
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(date_str.strip(), fmt).date()
            except (ValueError, AttributeError):
                continue
        return None

    def handle(self, *args, **options):
        school_id = options["school"]
        file_path = Path(options["file"])

        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        try:
            school = School.objects.get(pk=school_id)
        except School.DoesNotExist:
            raise CommandError(f"School with ID {school_id} not found")

        term = None
        if options.get("term"):
            term = AcademicTerm.objects.filter(
                pk=options["term"], school=school
            ).first()
        if not term:
            term = AcademicTerm.objects.filter(school=school, is_current=True).first()
        if not term:
            term = AcademicTerm.objects.filter(school=school).first()
        if not term:
            raise CommandError("No academic term found. Create one first.")

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # Auto-detect header row
        rows = list(ws.iter_rows(values_only=True))
        header_idx = None
        for i, row in enumerate(rows):
            vals = [str(v).strip() if v else "" for v in row]
            if "register" in vals or "register" in " ".join(vals).lower():
                header_idx = i
                break

        if header_idx is None:
            raise CommandError("Could not find header row containing 'Register'")

        headers = [str(v).strip().lower() if v else "" for v in rows[header_idx]]

        # Map columns by keyword matching
        col_map = {}
        for i, h in enumerate(headers):
            hl = h.lower()
            if "name" in hl or "surname" in hl:
                col_map["name"] = i
            elif "register" in hl or "matric" in hl or "id" in hl:
                col_map["register"] = i
            elif hl in ("sex", "gender"):
                col_map["sex"] = i
            elif "date" in hl and "birth" in hl:
                col_map["dob"] = i
            elif "place" in hl and "birth" in hl:
                col_map["pob"] = i
            elif "class" in hl or hl == "class":
                col_map["class"] = i
            elif "subclass" in hl or "sub" in hl:
                col_map["subclass"] = i
            elif "nationality" in hl:
                col_map["nationality"] = i

        required = ["name", "register", "sex", "dob"]
        missing = [r for r in required if r not in col_map]
        if missing:
            raise CommandError(
                f"Could not find columns: {', '.join(missing)}. Headers found: {headers}"
            )

        students_created = 0
        enrollments_created = 0
        skipped = 0
        errors = []

        for row in rows[header_idx + 1 :]:
            vals = [v for v in row]

            name_raw = (
                str(vals[col_map["name"]]).strip()
                if col_map["name"] < len(vals)
                else ""
            )
            reg_raw = (
                str(vals[col_map["register"]]).strip()
                if col_map["register"] < len(vals)
                else ""
            )
            sex_raw = (
                str(vals[col_map["sex"]]).strip().upper()
                if col_map["sex"] < len(vals)
                else ""
            )
            dob_raw = (
                str(vals[col_map["dob"]]).strip() if col_map["dob"] < len(vals) else ""
            )

            if not name_raw or not reg_raw:
                continue

            if sex_raw not in ("M", "F"):
                skipped += 1
                errors.append(f"  Invalid sex '{sex_raw}' for {name_raw}")
                continue

            dob = self._parse_date(dob_raw)
            if not dob:
                skipped += 1
                errors.append(f"  Invalid DOB '{dob_raw}' for {name_raw}")
                continue

            first_name, other_names = self._parse_name(name_raw)
            place_of_birth = (
                str(vals[col_map["pob"]]).strip()
                if col_map.get("pob") and col_map["pob"] < len(vals)
                else ""
            )

            # Lookup or create class
            class_name = (
                str(vals[col_map["class"]]).strip()
                if col_map.get("class") and col_map["class"] < len(vals)
                else ""
            )

            school_class = None
            if class_name:
                # Try to match by name or code
                school_class = SchoolClass.objects.filter(
                    school=school, name__iexact=class_name
                ).first()
                if not school_class:
                    school_class = SchoolClass.objects.filter(
                        school=school, code__iexact=class_name[:6]
                    ).first()

            try:
                student, created = Student.objects.get_or_create(
                    school=school,
                    unique_id=reg_raw,
                    defaults={
                        "first_name": first_name,
                        "other_names": other_names,
                        "sex": sex_raw,
                        "date_of_birth": dob,
                        "place_of_birth": place_of_birth,
                        "guardian_name": "",
                        "division_of_origin": "",
                        "region_of_origin": "",
                    },
                )
                if created:
                    students_created += 1

                if school_class:
                    enr, enr_created = StudentEnrollment.objects.get_or_create(
                        student=student,
                        academic_term=term,
                        defaults={"school_class": school_class},
                    )
                    if enr_created:
                        enrollments_created += 1
                    elif enr.school_class != school_class:
                        enr.school_class = school_class
                        enr.save(update_fields=["school_class"])

            except Exception as e:
                errors.append(f"  Error for {name_raw} ({reg_raw}): {e}")
                skipped += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Done. Created {students_created} students, "
                f"{enrollments_created} enrollments. "
                f"Skipped {skipped} rows."
            )
        )
        if errors:
            for e in errors[:5]:
                self.stdout.write(self.style.WARNING(e))
