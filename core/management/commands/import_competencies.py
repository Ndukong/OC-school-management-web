from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from core.models import AcademicTerm, Competency, School, Subject


class Command(BaseCommand):
    help = "Import subjects and competencies from the .xlsx template (all form levels)"

    def add_arguments(self, parser):
        parser.add_argument("--school", type=int, required=True, help="School ID")
        parser.add_argument("--file", type=str, required=True, help="Path to the .xlsx file")
        parser.add_argument("--year-start", type=int, default=2025)
        parser.add_argument("--year-end", type=int, default=2026)

    def handle(self, *args, **options):
        school_id = options["school"]
        file_path = Path(options["file"])
        year_start = options["year_start"]
        year_end = options["year_end"]

        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        try:
            school = School.objects.get(pk=school_id)
        except School.DoesNotExist:
            raise CommandError(f"School with ID {school_id} not found")

        term_map = {"FIRST": 1, "SECOND": 2, "THIRD": 3}
        terms = {}
        for tn in (1, 2, 3):
            t, _ = AcademicTerm.objects.get_or_create(
                school=school, term_number=tn,
                year_start=year_start, year_end=year_end,
            )
            terms[tn] = t

        wb = load_workbook(file_path, data_only=True)

        # ---- Sheet 1: Subjects reference ----
        ws = wb["Subjects"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = [v for v in row]
            code = str(vals[2]).strip() if vals[2] else ""
            name = str(vals[3]).strip() if vals[3] else ""
            if code and name:
                Subject.objects.get_or_create(
                    school=school, code=code,
                    defaults={"name": name},
                )

        # ---- Sheet 2: Competencies ----
        ws = wb["Competencies"]
        all_rows = list(ws.iter_rows(min_row=1, values_only=True))

        # Find all header rows ("Term" in column B)
        header_rows = []
        for idx, row in enumerate(all_rows):
            v = str(row[1]).strip() if row[1] else ""
            if v.upper() == "TERM":
                header_rows.append(idx)

        # 15 headers = 5 forms × 3 terms (in order: F1T1, F1T2, F1T3, F2T1, ...)
        if len(header_rows) != 15:
            self.stdout.write(self.style.WARNING(
                f"Expected 15 section headers, found {len(header_rows)}. Results may be incomplete."
            ))

        comp_count = 0
        for section_idx, hdr_idx in enumerate(header_rows):
            form_level = (section_idx // 3) + 1  # 0-2→F1, 3-5→F2, 6-8→F3, 9-11→F4, 12-14→F5
            term_offset = section_idx % 3       # 0=FIRST, 1=SECOND, 2=THIRD
            term_number = term_offset + 1

            # Data rows start after the header, end before next header or EOF
            next_hdr = header_rows[section_idx + 1] if section_idx + 1 < len(header_rows) else len(all_rows)
            data_rows = all_rows[hdr_idx + 1:next_hdr]

            prev = None
            for row in data_rows:
                term_raw = str(row[1]).strip().upper() if row[1] else ""
                subj_raw = str(row[2]).strip() if row[2] else ""

                if term_raw not in term_map or not subj_raw:
                    continue

                subj = Subject.objects.filter(school=school, name__iexact=subj_raw).first()
                if not subj:
                    subj = Subject.objects.filter(school=school, code__iexact=subj_raw).first()
                if not subj:
                    continue

                prev = None
                for ci in range(3, 7):
                    cv = str(row[ci]).strip() if ci < len(row) and row[ci] else ""
                    if not cv:
                        continue
                    if cv == '"' and prev:
                        cv = prev
                    Competency.objects.get_or_create(
                        subject=subj,
                        term=terms[term_number],
                        form_level=form_level,
                        sort_order=ci - 2,
                        defaults={"description": cv},
                    )
                    comp_count += 1
                    prev = cv

        self.stdout.write(self.style.SUCCESS(
            f"Done. Created/updated {comp_count} competencies for all form levels."
        ))
