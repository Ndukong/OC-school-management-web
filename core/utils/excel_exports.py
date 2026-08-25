
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill(start_color="E8F0F7", end_color="E8F0F7", fill_type="solid")
THIN = Side(style="thin", color="B8C6D4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_workbook(sheets: list[dict]) -> Workbook:
    """Build a styled workbook from sheet specs.

    Each sheet spec: {"title": str, "headers": list[str], "rows": list[list]}
    Rows may contain objects (str(date), Decimal, etc.) — values are written as-is.
    """
    wb = Workbook()
    wb.remove(wb.active)
    for spec in sheets:
        ws = wb.create_sheet(title=spec["title"])
        headers = spec["headers"]
        rows = spec.get("rows", [])
        freeze = spec.get("freeze", "A2")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
        ws.row_dimensions[1].height = 22

        for r, row in enumerate(rows, 2):
            for c, value in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=_to_excel(value))
                cell.border = BORDER
                if r % 2 == 0:
                    cell.fill = ALT_FILL

        for col, h in enumerate(headers, 1):
            width = max(
                10,
                min(40, max([len(str(h))] + [len(_cell_text(r[col - 1])) for r in rows]) + 2)
                if rows
                else len(str(h)) + 2,
            )
            ws.column_dimensions[get_column_letter(col)].width = width

        if rows:
            ws.freeze_panes = freeze
    return wb


def excel_response(wb: Workbook, filename: str) -> HttpResponse:
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _to_excel(value):
    if hasattr(value, "strftime") and not isinstance(value, (int, float)):
        return value
    return value


def _cell_text(value) -> str:
    if value is None:
        return ""
    return str(value)
