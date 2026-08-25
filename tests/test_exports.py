from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from django.contrib.auth.models import User
from django.test import Client
from django.urls import reverse
from openpyxl import load_workbook

from core.models import (
    AcademicTerm,
    AttendanceRecord,
    AttendanceRegister,
    ClassSubject,
    Competency,
    CompetencyScore,
    ExpenditureRecord,
    FeeType,
    IncomeRecord,
    School,
    SchoolClass,
    Student,
    StudentEnrollment,
    Subject,
    SubjectAverage,
    TermResult,
    UserProfile,
)
from tests.conftest import requires_pdf


@pytest.fixture
def data():
    school = School.objects.create(
        name_en="Export School",
        matricule="EXPORT1",
        region_en="North West",
        division_en="Mezam",
    )
    school_class = SchoolClass.objects.create(
        school=school, name="Form 2", code="F2", form_level=2, promotion_mark=8
    )
    term = AcademicTerm.objects.create(
        school=school, term_number=1, year_start=2025, year_end=2026, is_current=True
    )
    math = Subject.objects.create(school=school, name="Math", code="MAT")
    ClassSubject.objects.create(
        school_class=school_class, subject=math, coefficient=4, sort_order=1
    )
    comp = Competency.objects.create(
        subject=math, term=term, description="Term Exam", sort_order=1
    )
    s1 = Student.objects.create(
        school=school,
        first_name="Alice",
        sex="F",
        unique_id="111111111",
        date_of_birth=date(2010, 1, 1),
        place_of_birth="Bamenda",
        guardian_name="Guardian",
        division_of_origin="Mezam",
        region_of_origin="North West",
    )
    s2 = Student.objects.create(
        school=school,
        first_name="Bob",
        sex="M",
        unique_id="222222222",
        date_of_birth=date(2009, 5, 5),
        place_of_birth="Bamenda",
        guardian_name="Guardian",
        division_of_origin="Mezam",
        region_of_origin="North West",
    )
    for s in (s1, s2):
        StudentEnrollment.objects.create(
            student=s, school_class=school_class, academic_term=term
        )
    CompetencyScore.objects.create(
        student=s1, competency=comp, academic_term=term, score=Decimal(18)
    )
    CompetencyScore.objects.create(
        student=s2, competency=comp, academic_term=term, score=Decimal(12)
    )
    SubjectAverage.objects.create(
        student=s1, academic_term=term, subject=math,
        average=Decimal(18), grade="A", remark="Excellent",
    )
    SubjectAverage.objects.create(
        student=s2, academic_term=term, subject=math,
        average=Decimal(12), grade="C", remark="Good",
    )
    TermResult.objects.create(
        student=s1, academic_term=term, total_score=Decimal(72),
        total_coef=4, average=Decimal(18), rank=1, grade="A",
        remark="Excellent", promoted=True,
    )
    TermResult.objects.create(
        student=s2, academic_term=term, total_score=Decimal(48),
        total_coef=4, average=Decimal(12), rank=2, grade="C",
        remark="Good", promoted=False,
    )
    FeeType.objects.create(school=school, name="School Fees", category="PTA")
    IncomeRecord.objects.create(
        school=school, academic_term=term,
        fee_type=FeeType.objects.first(), amount=Decimal(1000),
        date_paid=date(2025, 10, 1), receipt_number="R1",
    )
    ExpenditureRecord.objects.create(
        school=school, academic_term=term, category="PTA",
        amount=Decimal(400), date=date(2025, 10, 5), description="Chairs",
    )
    register = AttendanceRegister.objects.create(
        school_class=school_class, date=date(2026, 1, 10), period=1
    )
    AttendanceRecord.objects.create(register=register, student=s1, status="P")
    AttendanceRecord.objects.create(register=register, student=s2, status="A")

    user = User.objects.create_user(username="admin2", password="pass", is_superuser=True)
    UserProfile.objects.create(user=user, school=school, role="admin")
    client = Client()
    client.login(username="admin2", password="pass")

    return {
        "school": school,
        "school_class": school_class,
        "term": term,
        "s1": s1,
        "s2": s2,
        "client": client,
    }


def assert_xlsx(content: bytes) -> dict:
    wb = load_workbook(BytesIO(content), data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        sheets[name] = [list(row) for row in ws.iter_rows(values_only=True)]
    return sheets


@pytest.mark.django_db
class TestExcelExports:
    def test_export_students(self, data):
        r = data["client"].get(reverse("export_students_excel"))
        assert r.status_code == 200
        assert r["Content-Type"].startswith("application/vnd.openxmlformats")
        sheets = assert_xlsx(r.content)
        headers = sheets["Students"][0]
        assert "Student ID" in headers and "Name" in headers
        names = [row[1] for row in sheets["Students"][1:]]
        assert "Alice" in names and "Bob" in names

    def test_export_marks(self, data):
        r = data["client"].get(
            reverse("export_marks_excel", args=[data["school_class"].id, data["term"].id])
        )
        assert r.status_code == 200
        sheets = assert_xlsx(r.content)
        assert sheets["Marks"][1][1] in ("Alice", "Bob")
        assert sheets["Marks"][1][4] == 18  # overall average of rank-1 student

    def test_export_finance(self, data):
        r = data["client"].get(reverse("export_finance_excel", args=[data["term"].id]))
        assert r.status_code == 200
        sheets = assert_xlsx(r.content)
        assert len(sheets["Income"]) == 2  # header + 1 income
        assert sheets["Summary"][3][1] == 1000  # total income

    def test_export_attendance(self, data):
        r = data["client"].get(
            reverse("export_attendance_excel", args=[data["school_class"].id, data["term"].id])
        )
        assert r.status_code == 200
        sheets = assert_xlsx(r.content)
        header = sheets["Attendance"][0]
        assert "Present" in header and "Absent" in header
        totals = {row[1]: row[2:] for row in sheets["Attendance"][1:]}
        assert totals["Alice"][0] == 1  # present count
        assert totals["Bob"][2] == 1  # absent count

    def test_export_results(self, data):
        r = data["client"].get(reverse("export_results_excel", args=[data["term"].id]))
        assert r.status_code == 200
        sheets = assert_xlsx(r.content)
        promoted = {row[2]: row[10] for row in sheets["Results"][1:]}
        assert promoted["Alice"] == "Promoted"
        assert promoted["Bob"] == "Repeat"


@pytest.mark.django_db
class TestNewReports:
    def test_reports_hub_renders(self, data):
        r = data["client"].get(reverse("reports_hub"))
        assert r.status_code == 200
        html = r.content.decode()
        assert "Report Cards" in html
        assert "Class Council Report" in html
        assert "PTA Financial Report" in html

    def test_class_council_renders_html(self, data):
        r = data["client"].get(
            reverse("class_council", args=[data["term"].id])
        )
        assert r.status_code == 200
        html = r.content.decode()
        assert "CLASS COUNCIL REPORT" in html
        assert "Form 2" in html
        assert "TOTAL" in html

    @requires_pdf
    def test_class_council_renders_pdf(self, data):
        r = data["client"].get(
            reverse("class_council", args=[data["term"].id]) + "?format=pdf"
        )
        assert r.status_code == 200
        assert r["Content-Type"] == "application/pdf"
        assert r.content[:4] == b"%PDF"

    def test_pta_financial_renders_html(self, data):
        r = data["client"].get(reverse("pta_financial", args=[data["term"].id]))
        assert r.status_code == 200
        html = r.content.decode()
        assert "PTA FINANCIAL REPORT" in html
        assert "1000" in html  # income amount

    @requires_pdf
    def test_pta_financial_renders_pdf(self, data):
        r = data["client"].get(
            reverse("pta_financial", args=[data["term"].id]) + "?format=pdf"
        )
        assert r.status_code == 200
        assert r["Content-Type"] == "application/pdf"
        assert r.content[:4] == b"%PDF"

    @requires_pdf
    def test_batch_report_cards(self, data):
        r = data["client"].get(
            reverse("batch_report_cards", args=[data["school_class"].id, data["term"].id])
        )
        assert r.status_code == 200
        assert r["Content-Type"] == "application/pdf"


def _blank_pdf_bytes() -> bytes:
    from io import BytesIO

    from pypdf import PdfWriter

    w = PdfWriter()
    w.add_blank_page(width=612, height=792)
    b = BytesIO()
    w.write(b)
    return b.getvalue()


@pytest.mark.django_db
class TestPdfBatchMerge:
    """Batch report card merge works with page-by-page copying (no WeasyPrint)."""

    def test_batch_report_cards_merge(self, data, monkeypatch):
        from core.utils.report_card import TermReportCard

        monkeypatch.setattr(
            TermReportCard, "render_pdf", lambda self, base_url=None: _blank_pdf_bytes()
        )
        r = data["client"].get(
            reverse("batch_report_cards", args=[data["school_class"].id, data["term"].id])
        )
        assert r.status_code == 200
        assert r["Content-Type"] == "application/pdf"
        assert r.content[:4] == b"%PDF"
        from io import BytesIO

        from pypdf import PdfReader

        assert len(PdfReader(BytesIO(r.content)).pages) == 2  # one per student

    def test_batch_annual_report_cards_merge(self, data, monkeypatch):
        from core.utils.report_card import AnnualReportCard

        # Create terms 2 & 3 plus enrollments and averages so students qualify.
        from core.models import AcademicTerm, StudentEnrollment, SubjectAverage

        d = data
        t2 = AcademicTerm.objects.create(
            school=d["school"], term_number=2, year_start=2025, year_end=2026
        )
        t3 = AcademicTerm.objects.create(
            school=d["school"], term_number=3, year_start=2025, year_end=2026
        )
        math = d["school_class"].subjects.first().subject
        for s in (d["s1"], d["s2"]):
            for t in (t2, t3):
                StudentEnrollment.objects.create(
                    student=s, school_class=d["school_class"], academic_term=t
                )
                SubjectAverage.objects.create(
                    student=s, academic_term=t, subject=math, average=Decimal(14)
                )

        monkeypatch.setattr(
            AnnualReportCard, "render_pdf", lambda self, base_url=None: _blank_pdf_bytes()
        )
        r = d["client"].get(
            reverse(
                "batch_annual_report_cards",
                args=[d["school_class"].id, 2025, 2026],
            )
        )
        assert r.status_code == 200
        assert r["Content-Type"] == "application/pdf"
        from io import BytesIO

        from pypdf import PdfReader

        assert len(PdfReader(BytesIO(r.content)).pages) == 2


@pytest.mark.django_db
class TestReportCountsDistinct:
    """JOIN duplication through enrollments must not inflate counts."""

    def _make_class_with_dupes(self, n_students=3, n_terms=3):
        school = School.objects.create(
            name_en="Dup School", matricule="DUP01",
            region_en="North West", division_en="Mezam",
        )
        cls = SchoolClass.objects.create(
            school=school, name="Form 1", code="F1", form_level=1, promotion_mark=8
        )
        terms = []
        for n in range(1, n_terms + 1):
            terms.append(
                AcademicTerm.objects.create(
                    school=school, term_number=n, year_start=2025, year_end=2026
                )
            )
        students = []
        for i in range(n_students):
            s = Student.objects.create(
                school=school,
                first_name=f"Student{i}",
                sex="M" if i % 2 == 0 else "F",
                unique_id=f"10000000{i}",
                date_of_birth=date(2010, 1, 1),
                place_of_birth="Bamenda",
                guardian_name="G",
                division_of_origin="Mezam",
                region_of_origin="North West",
            )
            students.append(s)
            for t in terms:
                StudentEnrollment.objects.create(
                    student=s, school_class=cls, academic_term=t
                )
                TermResult.objects.create(
                    student=s, academic_term=t,
                    average=Decimal(12 + i), total_score=Decimal(48 + i * 4),
                    total_coef=4, rank=i + 1,
                )
        return school, cls, terms[0], students

    def test_results_summary_counts_not_inflated(self):
        from core.utils.results_summary import ResultsSummary

        school, cls, term, _ = self._make_class_with_dupes()
        ctx = ResultsSummary(cls, term, school).get_context_data()
        assert ctx["num_sat"] == 3
        names = [st.full_name for _, st, _ in ctx["top3"]]
        assert len(set(names)) == 3
        assert ctx["enrolment_total"] == 3

    def test_class_council_term_counts_not_inflated(self):
        from core.utils.class_council import ClassCouncilReport

        school, cls, term, _ = self._make_class_with_dupes()
        ctx = ClassCouncilReport(term, school).get_context_data()
        f1 = [c for c in ctx["classes_data"] if c["school_class"].pk == cls.pk][0]
        assert f1["stats"]["on_roll_t"] == 3
        assert f1["stats"]["sat_t"] == 3
        assert f1["stats"]["passed_t"] == 3
        assert ctx["totals"]["on_roll_t"] == 3

    def test_pass_remark_thresholds(self):
        from core.utils.class_council import _pass_remark

        assert _pass_remark(Decimal(95)) == "Excellent"
        assert _pass_remark(Decimal(85)) == "Very good"
        assert _pass_remark(Decimal(75)) == "Good"
        assert _pass_remark(Decimal(65)) == "Fairly good"
        assert _pass_remark(Decimal(55)) == "Average"
        assert _pass_remark(Decimal(45)) == "Below average"
        assert _pass_remark(Decimal(35)) == "Weak"

    def test_results_summary_skips_empty_class(self):
        from core.utils.results_summary import ResultsSummary

        school, cls, term, _ = self._make_class_with_dupes()
        empty = SchoolClass.objects.create(
            school=school, name="Lower Sixth", code="LS", form_level=6
        )
        client = Client()
        user = User.objects.create_user(
            username="adminDup", password="pass", is_superuser=True
        )
        UserProfile.objects.create(user=user, school=school, role="admin")
        client.force_login(user)
        r = client.get(reverse("preview_results_summary", args=[term.id]))
        assert r.status_code == 200
        html = r.content.decode()
        assert "Lower Sixth" not in html
        assert "Form 1" in html
